import openpyxl
from datetime import date, timedelta
import json, os, re, glob

# ── 엑셀 파일 찾기 ───────────────────────────────────
xlsx_files = glob.glob("*.xlsx")
if not xlsx_files:
    raise FileNotFoundError("xlsx 파일을 찾을 수 없습니다.")
xlsx_path = xlsx_files[0]
print(f"파일 읽는 중: {xlsx_path}")

# ── 시트 찾기 (가장 최신 연도 or '2025') ────────────
wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
sheet_name = None
for name in wb.sheetnames:
    if re.match(r'20\d\d', name):
        sheet_name = name
if not sheet_name:
    sheet_name = wb.sheetnames[0]
print(f"시트: {sheet_name}")

ws = wb[sheet_name]
all_rows = list(ws.iter_rows(values_only=False))
dates      = [c.value for c in all_rows[0]]
subheaders = [c.value for c in all_rows[1]]

# ── 오늘 날짜 기준 ───────────────────────────────────
today           = date.today()
three_months_ago = date(today.year, today.month, today.day) - timedelta(days=90)

# ── 출고 열 / 재고 열 찾기 ──────────────────────────
outlet_cols = []
for i, (d, sub) in enumerate(zip(dates, subheaders)):
    if sub == '출고수량' and d is not None:
        d_date = d.date() if hasattr(d, 'date') else d
        if three_months_ago <= d_date <= today:
            outlet_cols.append(i)

stock_col = None
for i in range(len(dates)-1, -1, -1):
    d   = dates[i]
    sub = subheaders[i]
    if d is not None and sub == '재고수량':
        d_date = d.date() if hasattr(d, 'date') else d
        if d_date <= today:
            stock_col = i
            break

print(f"출고열 {len(outlet_cols)}개, 재고열 인덱스 {stock_col}")

# ── 품목 분석 ────────────────────────────────────────
results       = []
current_group = None

for row_cells in all_rows[2:]:
    cell_a = row_cells[0]
    name   = cell_a.value
    if not name or str(name).strip() == '':
        continue
    name = str(name).strip()
    bold = cell_a.font.bold if cell_a.font else False

    row_vals = [c.value for c in row_cells]
    stock    = row_vals[stock_col] if stock_col is not None else None
    if not isinstance(stock, (int, float)):
        stock = 0

    total_out = sum(
        row_vals[ci] for ci in outlet_cols
        if row_vals[ci] is not None and isinstance(row_vals[ci], (int, float))
    )
    monthly_avg = total_out / 3 if total_out else 0
    daily_avg   = monthly_avg / 30
    days_left   = round(stock / daily_avg) if daily_avg > 0 else 9999
    status      = 'danger' if days_left <= 45 else 'warning' if days_left <= 60 else 'ok'
    order_qty   = round(daily_avg * 90)
    deadline_days = days_left - 45
    deadline_str  = '이미 지남' if deadline_days <= 0 else (today + timedelta(days=deadline_days)).strftime('%Y-%m-%d')

    if bold:
        current_group = name
        if stock == 0 and monthly_avg == 0:
            continue
        results.append({
            'group': name, 'option': '', 'full_name': name,
            'stock': int(stock), 'monthly_avg': round(monthly_avg, 1),
            'days_left': days_left if days_left < 9999 else -1,
            'order_qty': order_qty, 'deadline_days': deadline_days,
            'deadline_str': deadline_str, 'status': status,
        })
    else:
        if stock == 0 and monthly_avg == 0:
            continue
        group     = current_group or ''
        full_name = f"{group} > {name}" if group else name
        results.append({
            'group': group, 'option': name, 'full_name': full_name,
            'stock': int(stock), 'monthly_avg': round(monthly_avg, 1),
            'days_left': days_left if days_left < 9999 else -1,
            'order_qty': order_qty, 'deadline_days': deadline_days,
            'deadline_str': deadline_str, 'status': status,
        })

danger  = len([r for r in results if r['status'] == 'danger'])
warning = len([r for r in results if r['status'] == 'warning'])
ok_c    = len([r for r in results if r['status'] == 'ok'])

summary = {
    'total': len(results), 'danger': danger,
    'warning': warning, 'ok': ok_c,
    'generated': today.strftime('%Y-%m-%d'),
    'filename': xlsx_path,
    'items': results,
}

json_str = json.dumps(summary, ensure_ascii=False, separators=(',',':'))
print(f"분석 완료: 전체 {len(results)}개, 즉시발주 {danger}개, 발주필요 {warning}개")

# ── index.html 읽기 및 데이터 교체 ──────────────────
with open('index.html', 'r', encoding='utf-8') as f:
    html = f.read()

# DATA = {...} 부분을 새 데이터로 교체
new_html = re.sub(
    r'const D = \{.*?\}(?=;)',
    f'const D = {json_str}',
    html, flags=re.DOTALL
)

# 기준일 업데이트
new_html = re.sub(
    r'기준일: <strong>.*?</strong>',
    f'기준일: <strong>{today.strftime("%Y-%m-%d")}</strong>',
    new_html
)

# 파일명 업데이트
new_html = re.sub(
    r'소스: .*?\.xlsx',
    f'소스: {xlsx_path}',
    new_html
)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(new_html)

print("index.html 업데이트 완료!")
